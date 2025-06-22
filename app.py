import pandas as pd
import random
import json
import os
import re
import shutil
from datetime import datetime, time
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from collections import defaultdict

class ScheduleGenerator:
    def __init__(self):
        self.lecturers = []
        self.subjects = []
        self.classes = []
        self.fixed_schedules = []  # Jadwal dari Excel
        self.generated_schedules = []  # Jadwal yang di-generate
        self.available_rooms = []
        self.break_times = [
            {"start": time(12, 0), "end": time(13, 0)},
            {"start": time(18, 0), "end": time(19, 0)}
        ]
        self.department_preferences = {
            "TI": [3, 4],
            "SI": [3, 4],
            "DKV": [5],
            "default": [3, 4, 5]
        }
        self.room_capacities = {}
        self.time_slots = [
            ("08:00", "09:40"), ("10:00", "11:40"), 
            ("13:00", "14:40"), ("15:00", "16:40"), 
            ("19:00", "20:40"), ("17:40 (online)", "19:20 (online)"),
            ("15:30 (online)", "17:10 (online)")
        ]
        self.excel_path = None  # Menyimpan path file Excel asli
        self.lecturer_breaks = defaultdict(list)  # Menyimpan waktu istirahat dosen

    def parse_time(self, time_str):
        try:
            time_str = str(time_str).strip()
            is_online = "(online)" in time_str.lower()
            time_part = re.sub(r'\(.*\)', '', time_str).strip()
            time_part = time_part.replace('.', ':')
            
            if ':' in time_part:
                hours, minutes = time_part.split(':')
                minutes = minutes[:2]
                time_part = f"{hours}:{minutes}"
            
            time_obj = datetime.strptime(time_part, "%H:%M").time()
            return time_obj, is_online
        except ValueError as e:
            print(f"Error parsing time '{time_str}': {e}")
            return None, False

    def is_valid_time_range(self, start_time_str, end_time_str):
        start_time, _ = self.parse_time(start_time_str)
        end_time, _ = self.parse_time(end_time_str)
        
        if not start_time or not end_time:
            return False
            
        if start_time >= end_time:
            return False
            
        return True

    def is_break_time(self, start_time_str, end_time_str):
        """Check if time range overlaps with break times"""
        start_time, _ = self.parse_time(start_time_str)
        end_time, _ = self.parse_time(end_time_str)
        if not start_time or not end_time:
            return False
            
        for bt in self.break_times:
            if start_time < bt['end'] and end_time > bt['start']:
                return True
        return False

    def load_data(self, excel_path):
        try:
            self.excel_path = excel_path  # Simpan path file asli
            df = pd.read_excel(excel_path, sheet_name='Mapping mata kuliah', skiprows=2)
            df = df.dropna(subset=['Nama Dosen', 'Mata Kuliah'])

            self.lecturers = df['Nama Dosen'].unique().tolist()
            self.subjects = df['Mata Kuliah'].unique().tolist()
            self.classes = df['Kelas'].unique().tolist()
            self.fixed_schedules = []
            
            for idx, row in df.iterrows():
                jam = row['Jam'] if pd.notna(row['Jam']) else ""
                hari = row['Hari'] if pd.notna(row['Hari']) else ""
                _, is_online = self.parse_time(jam)
                
                self.fixed_schedules.append({
                    'source': 'excel',  # Tandai berasal dari Excel
                    'excel_index': idx,  # Simpan indeks baris Excel
                    'dosen': row['Nama Dosen'],
                    'mata_kuliah': row['Mata Kuliah'],
                    'kelas': row['Kelas'],
                    'hari': hari,
                    'jam': jam,
                    'semester': row['Semester'],
                    'sks': row['SKS'],
                    'ruangan': 'Online' if is_online else row.get('Ruangan', ''),
                    'jumlah_mahasiswa': row.get('Jumlah Mahasiswa', 0)
                })
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Gagal memuat data: {str(e)}")
            return False

    def load_rooms(self, json_path):
        try:
            with open(json_path, 'r') as f:
                rooms = json.load(f)
                self.available_rooms = [room for room in rooms if 'online' not in room['nama'].lower()]
                self.room_capacities = {room['nama']: room.get('kapasitas', 30) for room in self.available_rooms}
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Gagal memuat data ruangan: {str(e)}")
            return False

    def load_rooms_from_excel(self, excel_path):
        try:
            df = pd.read_excel(excel_path)
            self.available_rooms = []
            for _, row in df.iterrows():
                room_name = row['Nama Ruangan']
                if 'online' not in str(room_name).lower():
                    self.available_rooms.append({
                        'nama': room_name,
                        'lantai': row.get('Lantai', 0),
                        'kapasitas': row.get('Kapasitas', 30)
                    })
            self.room_capacities = {room['nama']: room.get('kapasitas', 30) for room in self.available_rooms}
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Gagal memuat ruangan dari Excel: {str(e)}")
            return False

    def is_time_overlap(self, start1, end1, start2, end2):
        return not (end1 <= start2 or start1 >= end2)

    def is_conflict(self, schedule, check_room_capacity=True):
        try:
            if not schedule['jam']:  # Skip jika tidak ada jadwal
                return False
                
            # Handle case where jam is not in "start - end" format
            jam_parts = schedule['jam'].split(' - ')
            if len(jam_parts) != 2:
                return True
                
            start, end = jam_parts
            start_time, _ = self.parse_time(start)
            end_time, _ = self.parse_time(end)
            
            if not start_time or not end_time:
                return True
                
            if not self.is_valid_time_range(start, end):
                return True
                
            all_schedules = self.fixed_schedules + self.generated_schedules
            
            # 1. Check lecturer availability
            for sched in all_schedules:
                if sched != schedule and sched['dosen'] == schedule['dosen'] and sched['hari'] == schedule['hari'] and sched['jam']:
                    s_jam_parts = sched['jam'].split(' - ')
                    if len(s_jam_parts) != 2:
                        continue
                    s_start, s_end = s_jam_parts
                    s_start, _ = self.parse_time(s_start)
                    s_end, _ = self.parse_time(s_end)
                    if not s_start or not s_end:
                        continue
                    if self.is_time_overlap(start_time, end_time, s_start, s_end):
                        return True
                        
            # 2. Check room availability and capacity
            if check_room_capacity and schedule.get('ruangan') and schedule['ruangan'] != 'Online':
                # Room time conflict
                for sched in all_schedules:
                    if sched != schedule and sched.get('ruangan') == schedule['ruangan'] and sched['hari'] == schedule['hari'] and sched['jam']:
                        s_jam_parts = sched['jam'].split(' - ')
                        if len(s_jam_parts) != 2:
                            continue
                        s_start, s_end = s_jam_parts
                        s_start, _ = self.parse_time(s_start)
                        s_end, _ = self.parse_time(s_end)
                        if not s_start or not s_end:
                            continue
                        if self.is_time_overlap(start_time, end_time, s_start, s_end):
                            return True
                
                # Room capacity check
                room_capacity = self.room_capacities.get(schedule['ruangan'], 0)
                if schedule.get('jumlah_mahasiswa', 0) > room_capacity:
                    return True
            
            # 3. Check class availability (no same class at same time)
            for sched in all_schedules:
                if sched != schedule and sched['kelas'] == schedule['kelas'] and sched['hari'] == schedule['hari'] and sched['jam']:
                    s_jam_parts = sched['jam'].split(' - ')
                    if len(s_jam_parts) != 2:
                        continue
                    s_start, s_end = s_jam_parts
                    s_start, _ = self.parse_time(s_start)
                    s_end, _ = self.parse_time(s_end)
                    if not s_start or not s_end:
                        continue
                    if self.is_time_overlap(start_time, end_time, s_start, s_end):
                        return True
            
            # 4. Check break times
            if schedule.get('ruangan') != 'Online' and self.is_break_time(start, end):
                return True
                
            # 5. Check lecturer break times
            lecturer_breaks = self.lecturer_breaks.get(schedule['dosen'], [])
            for break_time in lecturer_breaks:
                break_start, break_end = break_time.split(' - ')
                break_start, _ = self.parse_time(break_start)
                break_end, _ = self.parse_time(break_end)
                
                if self.is_time_overlap(start_time, end_time, break_start, break_end):
                    return True
                
            return False
        except Exception as e:
            print(f"Error in conflict check: {e}")
            return True

    def get_available_room(self, department, day, start_time_str, end_time_str, student_count=0):
        try:
            start_time, is_online = self.parse_time(start_time_str)
            end_time, _ = self.parse_time(end_time_str)
            
            if is_online:
                return 'Online'
                
            if not start_time or not end_time:
                return None
                
            preferred_floors = self.department_preferences.get(department, self.department_preferences['default'])
            random.shuffle(self.available_rooms)

            for room in self.available_rooms:
                # Check capacity first
                if student_count > room.get('kapasitas', 30):
                    continue
                    
                room_floor = room.get('lantai')
                if room_floor not in preferred_floors:
                    continue

                available = True
                for sched in self.fixed_schedules + self.generated_schedules:
                    if sched.get('ruangan') == room['nama'] and sched['hari'] == day and sched['jam']:
                        try:
                            s_jam_parts = sched['jam'].split(' - ')
                            if len(s_jam_parts) != 2:
                                continue
                            s_start, s_end = s_jam_parts
                            s_start, _ = self.parse_time(s_start)
                            s_end, _ = self.parse_time(s_end)
                            if not s_start or not s_end:
                                continue
                                
                            if self.is_time_overlap(start_time, end_time, s_start, s_end):
                                available = False
                                break
                        except:
                            continue
                if available:
                    return room['nama']
            return None
        except Exception as e:
            print(f"Error in get_available_room: {e}")
            return None

    def generate_schedule_for_lecturer(self, lecturer_name):
        try:
            if not self.excel_path:
                messagebox.showerror("Error", "Tidak ada file Excel yang dimuat!")
                return False
                
            df = pd.read_excel(self.excel_path, sheet_name='Mapping mata kuliah', skiprows=2)
            df = df.dropna(subset=['Nama Dosen', 'Mata Kuliah'])
            unfixed = [
                {
                    'dosen': r['Nama Dosen'],
                    'mata_kuliah': r['Mata Kuliah'],
                    'kelas': r['Kelas'],
                    'semester': r['Semester'],
                    'sks': r['SKS'],
                    'jumlah_mahasiswa': r.get('Jumlah Mahasiswa', 0)
                }
                for _, r in df.iterrows()
                if r['Nama Dosen'] == lecturer_name and (pd.isna(r['Hari']) or pd.isna(r['Jam']))
            ]
            if not unfixed:
                print(f"Tidak ada jadwal kosong untuk dosen {lecturer_name}")
                return True
                
            days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
            success = 0
            
            for s in unfixed:
                for attempt in range(50):
                    day = random.choice(days)
                    start, end = random.choice(self.time_slots)
                    
                    is_online = "(online)" in start.lower() or "(online)" in end.lower()
                    
                    # Skip break times for all lecturers
                    if self.is_break_time(start, end):
                        continue
                        
                    temp_schedule = {
                        'source': 'generated',  # Tandai sebagai generated
                        'dosen': s['dosen'],
                        'mata_kuliah': s['mata_kuliah'],
                        'kelas': s['kelas'],
                        'hari': day,
                        'jam': f"{start} - {end}",
                        'semester': s['semester'],
                        'sks': s['sks'],
                        'jumlah_mahasiswa': s.get('jumlah_mahasiswa', 0)
                    }
                    
                    if self.is_conflict(temp_schedule, check_room_capacity=False):
                        continue
                        
                    if is_online:
                        room = 'Online'
                    else:
                        room = self.get_available_room(
                            s['kelas'][:2], 
                            day, 
                            start, 
                            end,
                            s.get('jumlah_mahasiswa', 0)
                        )
                        
                    if room:
                        temp_schedule['ruangan'] = room
                        if not self.is_conflict(temp_schedule):
                            # Tambahkan ke generated_schedules
                            self.generated_schedules.append(temp_schedule)
                            success += 1
                            break
            return success > 0
        except Exception as e:
            print(f"Error: {e}")
            return False

    def clear_all_rooms(self):
        for sched in self.fixed_schedules + self.generated_schedules:
            if '(online)' not in str(sched.get('jam', '')).lower():
                sched['ruangan'] = ''
        return True

    def fill_empty_rooms_randomly(self):
        try:
            self.clear_all_rooms()
            
            all_schedules = self.fixed_schedules + self.generated_schedules
            schedules_without_room = [
                s for s in all_schedules 
                if not s.get('ruangan') or str(s.get('ruangan')).strip() == ''
            ]
            
            random.shuffle(schedules_without_room)
            
            for sched in schedules_without_room:
                is_online = "(online)" in str(sched.get('jam', '')).lower()
                if is_online:
                    sched['ruangan'] = 'Online'
                    continue
                    
                if not sched.get('jam'):  # Skip jika tidak ada jadwal
                    continue
                    
                department = sched['kelas'][:2] if isinstance(sched['kelas'], str) and len(sched['kelas']) >= 2 else 'default'
                try:
                    jam_parts = sched['jam'].split(' - ')
                    if len(jam_parts) != 2:
                        continue
                    start, end = jam_parts
                    student_count = sched.get('jumlah_mahasiswa', 0)
                    
                    # Try preferred rooms first
                    room = self.get_available_room(
                        department, 
                        sched['hari'], 
                        start, 
                        end,
                        student_count
                    )
                    if room:
                        sched['ruangan'] = room
                        continue
                        
                    # Fallback to any available room
                    for room in self.available_rooms:
                        if student_count > room.get('kapasitas', 30):
                            continue
                            
                        available = True
                        for existing in all_schedules:
                            if existing.get('ruangan') == room['nama'] and existing['hari'] == sched['hari'] and existing.get('jam'):
                                try:
                                    e_jam_parts = existing['jam'].split(' - ')
                                    if len(e_jam_parts) != 2:
                                        continue
                                    e_start, e_end = e_jam_parts
                                    e_start, _ = self.parse_time(e_start)
                                    e_end, _ = self.parse_time(e_end)
                                    current_start, _ = self.parse_time(start)
                                    current_end, _ = self.parse_time(end)
                                    
                                    if self.is_time_overlap(current_start, current_end, e_start, e_end):
                                        available = False
                                        break
                                except:
                                    continue
                        if available:
                            sched['ruangan'] = room['nama']
                            break
                except Exception as e:
                    print(f"Error assigning room: {e}")
                    continue
            
            return True
        except Exception as e:
            print(f"Error in fill_empty_rooms_randomly: {e}")
            return False

    def save_to_excel(self, schedules, template_path, output_folder):
        try:
            output_path = os.path.join(output_folder, f"Jadwal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb = load_workbook(template_path)
            sheet = wb.active
            
            sheet.cell(row=3, column=1, value="Hari")
            sheet.cell(row=3, column=2, value="Mata Kuliah")
            sheet.cell(row=3, column=3, value="Kelas")
            sheet.cell(row=3, column=4, value="Ruangan")
            sheet.cell(row=3, column=5, value="Jam")
            sheet.cell(row=3, column=6, value="SKS")
            sheet.cell(row=3, column=7, value="Semester")
            sheet.cell(row=3, column=8, value="Dosen")
            sheet.cell(row=3, column=9, value="Jumlah Mahasiswa")
            
            row = 4
            for s in schedules:
                sheet.cell(row=row, column=1, value=s['hari'])
                sheet.cell(row=row, column=2, value=s['mata_kuliah'])
                sheet.cell(row=row, column=3, value=s['kelas'])
                sheet.cell(row=row, column=4, value=s.get('ruangan', ''))
                sheet.cell(row=row, column=5, value=s['jam'])
                sheet.cell(row=row, column=6, value=s['sks'])
                sheet.cell(row=row, column=7, value=s['semester'])
                sheet.cell(row=row, column=8, value=s['dosen'])
                sheet.cell(row=row, column=9, value=s.get('jumlah_mahasiswa', ''))
                row += 1
            wb.save(output_path)
            return output_path
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan: {str(e)}")
            return None

    def update_excel_file(self, schedule, new_schedule):
        """Memperbarui file Excel asli dengan perubahan jadwal"""
        try:
            if not self.excel_path:
                messagebox.showerror("Error", "Tidak ada file Excel yang dimuat")
                return False
                
            # Buat backup file asli
            backup_path = self.excel_path.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            shutil.copyfile(self.excel_path, backup_path)
            
            # Baca file Excel
            wb = load_workbook(self.excel_path)
            sheet = wb['Mapping mata kuliah']
            
            # Temukan baris yang sesuai (indeks dimulai dari 1)
            row_index = schedule['excel_index'] + 4  # Skip 2 header + 2 baris kosong
            
            # Perbarui nilai di Excel
            sheet.cell(row=row_index, column=1, value=new_schedule['hari'])  # Kolom Hari
            sheet.cell(row=row_index, column=8, value=new_schedule['jam'])    # Kolom Jam
            sheet.cell(row=row_index, column=7, value=new_schedule.get('ruangan', ''))  # Kolom Ruangan
            # Jika ada perubahan pada data lain, tambahkan di sini
            
            # Simpan perubahan
            wb.save(self.excel_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Gagal memperbarui file Excel: {str(e)}")
            return False

    def find_all_conflicts(self):
        conflicts = {
            'lecturer': [],
            'room': [],
            'class': [],
            'capacity': [],
            'empty_room': [],
            'break_time': []  # Konflik waktu istirahat
        }
        
        all_schedules = self.fixed_schedules + self.generated_schedules
        
        for i, sched in enumerate(all_schedules):
            if not sched.get('jam'):  # Skip jika tidak ada jadwal
                continue
                
            # Lecturer conflicts
            for j, other in enumerate(all_schedules[i+1:], i+1):
                if sched['dosen'] == other['dosen'] and sched['hari'] == other['hari'] and other.get('jam'):
                    try:
                        s_jam_parts = sched['jam'].split(' - ')
                        o_jam_parts = other['jam'].split(' - ')
                        if len(s_jam_parts) != 2 or len(o_jam_parts) != 2:
                            continue
                            
                        s_start, s_end = s_jam_parts
                        o_start, o_end = o_jam_parts
                        
                        s_start, _ = self.parse_time(s_start)
                        s_end, _ = self.parse_time(s_end)
                        o_start, _ = self.parse_time(o_start)
                        o_end, _ = self.parse_time(o_end)
                        
                        if not s_start or not s_end or not o_start or not o_end:
                            continue
                            
                        if self.is_time_overlap(s_start, s_end, o_start, o_end):
                            conflicts['lecturer'].append({
                                'conflict_type': 'Dosen ganda',
                                'dosen': sched['dosen'],
                                'hari': sched['hari'],
                                'waktu': f"{max(s_start, o_start)}-{min(s_end, o_end)}",
                                'schedule1': sched,
                                'schedule2': other
                            })
                    except:
                        continue
            
            # Room conflicts and capacity issues
            if sched.get('ruangan') and sched['ruangan'] != 'Online' and sched.get('jam'):
                # Room time conflicts
                for j, other in enumerate(all_schedules[i+1:], i+1):
                    if other.get('ruangan') == sched['ruangan'] and other['hari'] == sched['hari'] and other.get('jam'):
                        try:
                            s_jam_parts = sched['jam'].split(' - ')
                            o_jam_parts = other['jam'].split(' - ')
                            if len(s_jam_parts) != 2 or len(o_jam_parts) != 2:
                                continue
                                
                            s_start, s_end = s_jam_parts
                            o_start, o_end = o_jam_parts
                            
                            s_start, _ = self.parse_time(s_start)
                            s_end, _ = self.parse_time(s_end)
                            o_start, _ = self.parse_time(o_start)
                            o_end, _ = self.parse_time(o_end)
                            
                            if not s_start or not s_end or not o_start or not o_end:
                                continue
                                
                            if self.is_time_overlap(s_start, s_end, o_start, o_end):
                                conflicts['room'].append({
                                    'conflict_type': 'Ruangan ganda',
                                    'ruangan': sched['ruangan'],
                                    'hari': sched['hari'],
                                    'waktu': f"{max(s_start, o_start)}-{min(s_end, o_end)}",
                                    'schedule1': sched,
                                    'schedule2': other
                                })
                        except:
                            continue
                
                # Room capacity issues
                room_capacity = self.room_capacities.get(sched['ruangan'], 0)
                if sched.get('jumlah_mahasiswa', 0) > room_capacity:
                    conflicts['capacity'].append({
                        'conflict_type': 'Kapasitas ruangan terlampaui',
                        'ruangan': sched['ruangan'],
                        'kapasitas': room_capacity,
                        'mahasiswa': sched.get('jumlah_mahasiswa', 0),
                        'schedule': sched
                    })
            
            # Class conflicts
            for j, other in enumerate(all_schedules[i+1:], i+1):
                if sched['kelas'] == other['kelas'] and sched['hari'] == other['hari'] and other.get('jam'):
                    try:
                        s_jam_parts = sched['jam'].split(' - ')
                        o_jam_parts = other['jam'].split(' - ')
                        if len(s_jam_parts) != 2 or len(o_jam_parts) != 2:
                            continue
                            
                        s_start, s_end = s_jam_parts
                        o_start, o_end = o_jam_parts
                        
                        s_start, _ = self.parse_time(s_start)
                        s_end, _ = self.parse_time(s_end)
                        o_start, _ = self.parse_time(o_start)
                        o_end, _ = self.parse_time(o_end)
                        
                        if not s_start or not s_end or not o_start or not o_end:
                            continue
                            
                        if self.is_time_overlap(s_start, s_end, o_start, o_end):
                            conflicts['class'].append({
                                'conflict_type': 'Kelas ganda',
                                'kelas': sched['kelas'],
                                'hari': sched['hari'],
                                'waktu': f"{max(s_start, o_start)}-{min(s_end, o_end)}",
                                'schedule1': sched,
                                'schedule2': other
                            })
                    except:
                        continue
            
            # Break time conflicts
            if sched.get('ruangan') != 'Online' and sched.get('jam'):
                try:
                    jam_parts = sched['jam'].split(' - ')
                    if len(jam_parts) == 2:
                        start, end = jam_parts
                        if self.is_break_time(start, end):
                            conflicts['break_time'].append({
                                'conflict_type': 'Waktu istirahat',
                                'dosen': sched['dosen'],
                                'hari': sched['hari'],
                                'waktu': sched['jam'],
                                'schedule': sched
                            })
                except:
                    pass
        
        # Check for empty rooms
        days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
        
        for room in self.available_rooms:
            room_name = room['nama']
            
            for day in days:
                for start, end in self.time_slots[:5]:  # Skip online slots
                    occupied = False
                    for sched in all_schedules:
                        if sched.get('ruangan') == room_name and sched['hari'] == day and sched.get('jam'):
                            try:
                                s_jam_parts = sched['jam'].split(' - ')
                                if len(s_jam_parts) != 2:
                                    continue
                                s_start, s_end = s_jam_parts
                                s_start, _ = self.parse_time(s_start)
                                s_end, _ = self.parse_time(s_end)
                                current_start, _ = self.parse_time(start)
                                current_end, _ = self.parse_time(end)
                                
                                if not s_start or not s_end:
                                    continue
                                    
                                if self.is_time_overlap(current_start, current_end, s_start, s_end):
                                    occupied = True
                                    break
                            except:
                                continue
                    
                    if not occupied:
                        conflicts['empty_room'].append({
                            'conflict_type': 'Ruangan kosong',
                            'ruangan': room_name,
                            'hari': day,
                            'waktu': f"{start} - {end}",
                            'lantai': room.get('lantai', '?'),
                            'kapasitas': room.get('kapasitas', '?')
                        })
        
        return conflicts
    
    def suggest_conflict_resolutions(self, conflict):
        suggestions = []
        
        if conflict['conflict_type'] == 'Dosen ganda':
            days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
            other_days = [d for d in days if d != conflict['hari']]
            
            for day in other_days:
                temp_schedule = conflict['schedule1'].copy()
                temp_schedule['hari'] = day
                if not self.is_conflict(temp_schedule):
                    suggestions.append(f"Pindahkan {temp_schedule['mata_kuliah']} ke hari {day}")
                    break
            
            for slot in self.time_slots:
                temp_schedule = conflict['schedule1'].copy()
                temp_schedule['jam'] = f"{slot[0]} - {slot[1]}"
                if not self.is_conflict(temp_schedule):
                    suggestions.append(f"Ubah jam {temp_schedule['mata_kuliah']} menjadi {slot[0]}-{slot[1]}")
                    break
        
        elif conflict['conflict_type'] == 'Ruangan ganda':
            department = conflict['schedule1']['kelas'][:2] if len(conflict['schedule1']['kelas']) >= 2 else 'default'
            jam_parts = conflict['schedule1']['jam'].split(' - ')
            if len(jam_parts) != 2:
                suggestions.append("Format waktu tidak valid")
                return suggestions
                
            start, end = jam_parts
            student_count = conflict['schedule1'].get('jumlah_mahasiswa', 0)
            
            # Try preferred rooms first
            alt_room = self.get_available_room(
                department,
                conflict['hari'],
                start,
                end,
                student_count
            )
            
            if alt_room and alt_room != conflict['ruangan']:
                suggestions.append(f"Ganti ruangan {conflict['ruangan']} dengan {alt_room}")
            
            suggestions.append("Ubah salah satu kelas menjadi online")
        
        elif conflict['conflict_type'] == 'Kelas ganda':
            suggestions.append("Ubah salah satu kelas menjadi online")
            
            days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
            other_days = [d for d in days if d != conflict['hari']]
            
            for day in other_days:
                temp_schedule = conflict['schedule1'].copy()
                temp_schedule['hari'] = day
                if not self.is_conflict(temp_schedule):
                    suggestions.append(f"Pindahkan {temp_schedule['mata_kuliah']} ke hari {day}")
                    break
        
        elif conflict['conflict_type'] == 'Kapasitas ruangan terlampaui':
            required_capacity = conflict['mahasiswa']
            current_room = conflict['ruangan']
            
            for room in self.available_rooms:
                if room['nama'] != current_room and room.get('kapasitas', 0) >= required_capacity:
                    available = True
                    for sched in self.fixed_schedules + self.generated_schedules:
                        if sched.get('ruangan') == room['nama'] and sched['hari'] == conflict['schedule']['hari'] and sched.get('jam'):
                            try:
                                s_jam_parts = sched['jam'].split(' - ')
                                if len(s_jam_parts) != 2:
                                    continue
                                s_start, s_end = s_jam_parts
                                s_start, _ = self.parse_time(s_start)
                                s_end, _ = self.parse_time(s_end)
                                c_jam_parts = conflict['schedule']['jam'].split(' - ')
                                if len(c_jam_parts) != 2:
                                    continue
                                c_start, c_end = c_jam_parts
                                c_start, _ = self.parse_time(c_start)
                                c_end, _ = self.parse_time(c_end)
                                
                                if not s_start or not s_end or not c_start or not c_end:
                                    continue
                                    
                                if self.is_time_overlap(c_start, c_end, s_start, s_end):
                                    available = False
                                    break
                            except:
                                continue
                    if available:
                        suggestions.append(f"Ganti ruangan {current_room} dengan {room['nama']} (kapasitas: {room['kapasitas']})")
                        break
            
            suggestions.append("Pindahkan ke ruangan dengan kapasitas lebih besar")
            suggestions.append("Pisahkan kelas menjadi dua sesi")
        
        elif conflict['conflict_type'] == 'Ruangan kosong':
            department = None
            for dept, floors in self.department_preferences.items():
                if dept != 'default' and conflict['lantai'] in floors:
                    department = dept
                    break
            
            if department:
                suggestions.append(f"Bisa digunakan untuk kelas {department} (lantai {conflict['lantai']})")
            else:
                suggestions.append(f"Ruangan tersedia di lantai {conflict['lantai']} (kapasitas: {conflict['kapasitas']})")
            
            suggestions.append("Bisa digunakan untuk make-up class")
            suggestions.append("Bisa digunakan untuk rapat atau kegiatan lain")
        
        elif conflict['conflict_type'] == 'Waktu istirahat':
            suggestions.append("Pindahkan ke waktu sebelum pukul 12:00 atau setelah pukul 13:00")
            suggestions.append("Pindahkan ke waktu sebelum pukul 18:00 atau setelah pukul 19:00")
            suggestions.append("Ubah menjadi kelas online")
        
        if not suggestions:
            suggestions.append("Tidak ada solusi otomatis tersedia. Perlu penyesuaian manual.")
        
        return suggestions

    def add_manual_schedule(self, schedule):
        # Untuk manual, tambahkan sebagai fixed schedule
        schedule['source'] = 'manual'
        self.fixed_schedules.append(schedule)
        
        # Update lists if new entries
        if schedule['dosen'] not in self.lecturers:
            self.lecturers.append(schedule['dosen'])
        if schedule['mata_kuliah'] not in self.subjects:
            self.subjects.append(schedule['mata_kuliah'])
        if schedule['kelas'] not in self.classes:
            self.classes.append(schedule['kelas'])
            
        return True

    def remove_schedule(self, schedule):
        if schedule in self.fixed_schedules:
            self.fixed_schedules.remove(schedule)
            return True
        elif schedule in self.generated_schedules:
            self.generated_schedules.remove(schedule)
            return True
        return False

    def edit_schedule(self, old_schedule, new_schedule):
        # Jika berasal dari Excel, perbarui file Excel
        if old_schedule.get('source') == 'excel':
            if not self.update_excel_file(old_schedule, new_schedule):
                return False
        
        # Perbarui data di memori
        if self.remove_schedule(old_schedule):
            # Pertahankan source
            new_schedule['source'] = old_schedule.get('source', 'manual')
            
            # Untuk jadwal Excel, pertahankan excel_index
            if old_schedule.get('source') == 'excel':
                new_schedule['excel_index'] = old_schedule['excel_index']
            
            # Tambahkan jadwal baru
            if new_schedule.get('source') == 'excel':
                self.fixed_schedules.append(new_schedule)
            else:
                self.add_manual_schedule(new_schedule)
                
            return True
        return False

    def auto_resolve_conflicts(self):
        """Fungsi untuk menyelesaikan konflik secara otomatis"""
        resolved = 0
        conflicts = self.find_all_conflicts()
        
        # Resolve lecturer conflicts
        for conflict in conflicts['lecturer']:
            # Coba pindahkan jadwal pertama ke hari lain
            for day in ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']:
                if day == conflict['schedule1']['hari']:
                    continue
                    
                new_schedule = conflict['schedule1'].copy()
                new_schedule['hari'] = day
                
                if not self.is_conflict(new_schedule):
                    if self.edit_schedule(conflict['schedule1'], new_schedule):
                        resolved += 1
                        break
                        
        return resolved

    def add_lecturer_break(self, lecturer, day, start_time, end_time):
        """Menambahkan waktu istirahat untuk dosen tertentu"""
        key = f"{lecturer}|{day}"
        self.lecturer_breaks[key].append(f"{start_time} - {end_time}")


class ManualInputDialog(tk.Toplevel):
    def __init__(self, parent, generator, callback, schedule=None):
        super().__init__(parent)
        if schedule:
            self.title("Edit Jadwal")
        else:
            self.title("Tambah Jadwal Manual")
        self.generator = generator
        self.callback = callback
        self.schedule = schedule
        
        # Form fields
        ttk.Label(self, text="Hari:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.hari_var = tk.StringVar()
        hari_options = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
        ttk.OptionMenu(self, self.hari_var, hari_options[0], *hari_options).grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Dosen:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.dosen_var = tk.StringVar()
        self.dosen_entry = ttk.Combobox(self, textvariable=self.dosen_var)
        self.dosen_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Mata Kuliah:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.matkul_var = tk.StringVar()
        self.matkul_entry = ttk.Entry(self, textvariable=self.matkul_var)
        self.matkul_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Kelas:").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        self.kelas_var = tk.StringVar()
        self.kelas_entry = ttk.Entry(self, textvariable=self.kelas_var)
        self.kelas_entry.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Ruangan:").grid(row=4, column=0, padx=5, pady=5, sticky='e')
        self.ruangan_var = tk.StringVar()
        self.ruangan_entry = ttk.Combobox(self, textvariable=self.ruangan_var)
        self.ruangan_entry.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Jam (HH:MM - HH:MM):").grid(row=5, column=0, padx=5, pady=5, sticky='e')
        self.jam_var = tk.StringVar()
        self.jam_entry = ttk.Entry(self, textvariable=self.jam_var)
        self.jam_entry.grid(row=5, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="SKS:").grid(row=6, column=0, padx=5, pady=5, sticky='e')
        self.sks_var = tk.StringVar()
        self.sks_entry = ttk.Entry(self, textvariable=self.sks_var)
        self.sks_entry.grid(row=6, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Semester:").grid(row=7, column=0, padx=5, pady=5, sticky='e')
        self.semester_var = tk.StringVar()
        self.semester_entry = ttk.Entry(self, textvariable=self.semester_var)
        self.semester_entry.grid(row=7, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Jumlah Mahasiswa:").grid(row=8, column=0, padx=5, pady=5, sticky='e')
        self.mahasiswa_var = tk.StringVar()
        self.mahasiswa_entry = ttk.Entry(self, textvariable=self.mahasiswa_var)
        self.mahasiswa_entry.grid(row=8, column=1, padx=5, pady=5, sticky='w')
        
        # Buttons
        if schedule:
            button_text = "Simpan Perubahan"
        else:
            button_text = "Tambah"
        ttk.Button(self, text=button_text, command=self.save_schedule).grid(row=9, column=0, columnspan=2, pady=10)
        
        # Populate dropdowns
        self.update_dropdowns()
        
        # If editing, populate the fields
        if schedule:
            self.hari_var.set(schedule['hari'])
            self.dosen_var.set(schedule['dosen'])
            self.matkul_var.set(schedule['mata_kuliah'])
            self.kelas_var.set(schedule['kelas'])
            self.ruangan_var.set(schedule.get('ruangan', ''))
            self.jam_var.set(schedule['jam'])
            self.sks_var.set(str(schedule['sks']))
            self.semester_var.set(str(schedule['semester']))
            self.mahasiswa_var.set(str(schedule.get('jumlah_mahasiswa', 0)))
        
    def update_dropdowns(self):
        # Update dosen dropdown
        self.dosen_entry['values'] = self.generator.lecturers
        
        # Update ruangan dropdown
        room_names = [room['nama'] for room in self.generator.available_rooms]
        room_names.append('Online')
        self.ruangan_entry['values'] = room_names
        
    def save_schedule(self):
        try:
            new_schedule = {
                'dosen': self.dosen_var.get(),
                'mata_kuliah': self.matkul_var.get(),
                'kelas': self.kelas_var.get(),
                'hari': self.hari_var.get(),
                'jam': self.jam_var.get(),
                'semester': int(self.semester_var.get() or 0),
                'sks': int(self.sks_var.get() or 0),
                'ruangan': self.ruangan_var.get(),
                'jumlah_mahasiswa': int(self.mahasiswa_var.get() or 0)
            }
            
            # Validate required fields
            if not all([new_schedule['dosen'], new_schedule['mata_kuliah'], new_schedule['kelas']]):
                messagebox.showerror("Error", "Dosen, Mata Kuliah, dan Kelas harus diisi!")
                return
                
            # Validate time format if jam is provided
            if new_schedule['jam']:
                parts = new_schedule['jam'].split(' - ')
                if len(parts) != 2:
                    messagebox.showerror("Error", "Format waktu tidak valid! Gunakan format 'HH:MM - HH:MM'")
                    return
                    
                if not self.generator.is_valid_time_range(parts[0], parts[1]):
                    messagebox.showerror("Error", "Format waktu tidak valid! Pastikan format HH:MM - HH:MM dan waktu mulai sebelum waktu selesai.")
                    return
                
            if self.schedule:
                # Editing existing schedule
                if self.generator.edit_schedule(self.schedule, new_schedule):
                    messagebox.showinfo("Sukses", "Jadwal berhasil diperbarui!")
                else:
                    messagebox.showerror("Error", "Gagal memperbarui jadwal")
            else:
                # Adding new schedule
                self.generator.add_manual_schedule(new_schedule)
                messagebox.showinfo("Sukses", "Jadwal berhasil ditambahkan!")
            
            self.callback()
            self.destroy()
            
        except ValueError:
            messagebox.showerror("Error", "Pastikan SKS, Semester, dan Jumlah Mahasiswa berupa angka!")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan jadwal: {str(e)}")


class BreakTimeDialog(tk.Toplevel):
    """Dialog untuk menambahkan waktu istirahat dosen"""
    def __init__(self, parent, generator, callback):
        super().__init__(parent)
        self.title("Tambah Waktu Istirahat Dosen")
        self.generator = generator
        self.callback = callback
        
        # Form fields
        ttk.Label(self, text="Dosen:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.dosen_var = tk.StringVar()
        self.dosen_entry = ttk.Combobox(self, textvariable=self.dosen_var)
        self.dosen_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Hari:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.hari_var = tk.StringVar()
        hari_options = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
        ttk.OptionMenu(self, self.hari_var, hari_options[0], *hari_options).grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Waktu Mulai (HH:MM):").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.start_var = tk.StringVar(value="12:00")
        self.start_entry = ttk.Entry(self, textvariable=self.start_var)
        self.start_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Waktu Selesai (HH:MM):").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        self.end_var = tk.StringVar(value="13:00")
        self.end_entry = ttk.Entry(self, textvariable=self.end_var)
        self.end_entry.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        # Buttons
        ttk.Button(self, text="Tambahkan", command=self.add_break).grid(row=4, column=0, columnspan=2, pady=10)
        
        # Populate dropdowns
        self.update_dropdowns()
        
    def update_dropdowns(self):
        # Update dosen dropdown
        self.dosen_entry['values'] = self.generator.lecturers
        
    def add_break(self):
        try:
            dosen = self.dosen_var.get()
            hari = self.hari_var.get()
            start = self.start_var.get()
            end = self.end_var.get()
            
            if not dosen:
                messagebox.showerror("Error", "Pilih dosen terlebih dahulu!")
                return
                
            # Validasi format waktu
            if not re.match(r'^\d{1,2}:\d{2}$', start) or not re.match(r'^\d{1,2}:\d{2}$', end):
                messagebox.showerror("Error", "Format waktu tidak valid! Gunakan format HH:MM")
                return
                
            # Tambahkan waktu istirahat
            self.generator.add_lecturer_break(dosen, hari, start, end)
            messagebox.showinfo("Sukses", f"Waktu istirahat berhasil ditambahkan untuk {dosen} pada hari {hari} ({start} - {end})")
            self.destroy()
            self.callback()
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menambahkan waktu istirahat: {str(e)}")


class ScheduleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Nusaputra Schedule Generator")
        self.root.geometry("1000x800")
        self.generator = ScheduleGenerator()
        self.generator.load_rooms("data/rooms.json")
        self.sort_order_hari = 'asc'
        self.current_filter_hari = None
        self.selected_schedule = None
        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(title_frame, 
                 text="Penjadwalan Kuliah Universitas Nusaputra", 
                 font=("Arial", 16, "bold")).pack()
        
        ttk.Label(title_frame,
                 text="Versi 5.0 - Dengan waktu istirahat dosen",
                 font=("Arial", 10)).pack()

        filter_frame = ttk.Frame(main_frame)
        filter_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(filter_frame, text="Filter Hari:").pack(side=tk.LEFT)
        self.hari_var = tk.StringVar()
        self.hari_dropdown = ttk.Combobox(filter_frame, 
                                        textvariable=self.hari_var,
                                        values=['Semua', 'Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat'],
                                        state='readonly')
        self.hari_dropdown.pack(side=tk.LEFT, padx=5)
        self.hari_dropdown.set('Semua')
        self.hari_dropdown.bind("<<ComboboxSelected>>", self.apply_filters)
        
        self.sort_hari_btn = ttk.Button(filter_frame, 
                                      text="Sort Hari (A-Z)", 
                                      command=self.toggle_sort_hari)
        self.sort_hari_btn.pack(side=tk.LEFT, padx=5)

        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(control_frame, text="Pilih Dosen:").pack(side=tk.LEFT)
        self.lecturer_var = tk.StringVar()
        self.lecturer_dropdown = ttk.Combobox(control_frame, 
                                             textvariable=self.lecturer_var, 
                                             width=40)
        self.lecturer_dropdown.pack(side=tk.LEFT, padx=5)
        self.lecturer_dropdown.bind("<<ComboboxSelected>>", self.show_lecturer_schedule)
        
        ttk.Button(control_frame, 
                  text="Load Excel", 
                  command=self.load_excel_data).pack(side=tk.RIGHT, padx=5)
        
        ttk.Button(control_frame,
                  text="Load Ruangan (JSON)",
                  command=self.load_room_data_json).pack(side=tk.RIGHT, padx=5)

        ttk.Button(control_frame,
                  text="Load Ruangan (Excel)",
                  command=self.load_room_data_excel).pack(side=tk.RIGHT, padx=5)

        schedule_frame = ttk.Frame(main_frame)
        schedule_frame.pack(fill=tk.BOTH, expand=True)
        
        self.schedule_tree = ttk.Treeview(schedule_frame, 
                                        columns=('Hari', 'Mata Kuliah', 'Kelas', 'Ruangan', 'Kapasitas', 'Jam', 'SKS', 'Semester', 'Dosen', 'Mahasiswa'), 
                                        show='headings',
                                        height=12)
        
        columns = [
            ('Hari', 80), ('Mata Kuliah', 150), ('Kelas', 60), 
            ('Ruangan', 80), ('Kapasitas', 70), ('Jam', 120),
            ('SKS', 40), ('Semester', 70), ('Dosen', 120), ('Mahasiswa', 70)
        ]
        
        for col, width in columns:
            self.schedule_tree.heading(col, text=col)
            self.schedule_tree.column(col, width=width, anchor='center')
        
        scroll_y = ttk.Scrollbar(schedule_frame, orient="vertical", command=self.schedule_tree.yview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.schedule_tree.configure(yscrollcommand=scroll_y.set)
        self.schedule_tree.pack(fill=tk.BOTH, expand=True)

        # Bind selection event
        self.schedule_tree.bind('<<TreeviewSelect>>', self.on_schedule_select)

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=5)
        
        buttons = [
            ("Generate Dosen Ini", self.generate_for_lecturer),
            ("Hapus Ruangan", self.clear_rooms),
            ("Generate Ruangan", self.generate_rooms),
            ("Simpan Semua", self.save_schedule_all),
            ("Simpan Dosen Ini", self.save_schedule_for_current_lecturer),
            ("Cek Konflik", self.show_conflicts),
            ("Tambah Manual", self.show_manual_input),
            ("Edit Jadwal", self.edit_selected_schedule),
            ("Hapus Jadwal", self.delete_selected_schedule),
            ("Atasi Konflik", self.resolve_conflicts),
            ("Simpan ke Excel Asli", self.save_to_original_excel),
            ("Tambah Istirahat", self.add_break_time)  # Tombol baru untuk waktu istirahat
        ]
        
        # Create buttons in two rows to avoid overflow
        top_button_frame = ttk.Frame(button_frame)
        top_button_frame.pack(fill=tk.X, pady=2)
        
        bottom_button_frame = ttk.Frame(button_frame)
        bottom_button_frame.pack(fill=tk.X, pady=2)
        
        top_buttons = buttons[:6]
        bottom_buttons = buttons[6:]
        
        for text, command in top_buttons:
            ttk.Button(top_button_frame, 
                      text=text, 
                      command=command).pack(side=tk.LEFT, padx=2)
                      
        for text, command in bottom_buttons:
            ttk.Button(bottom_button_frame, 
                      text=text, 
                      command=command).pack(side=tk.LEFT, padx=2)

        self.conflict_frame = ttk.Frame(main_frame)
        
        self.conflict_tree = ttk.Treeview(self.conflict_frame, 
                                        columns=('Tipe', 'Entitas', 'Hari', 'Waktu', 'Detail1', 'Detail2', 'Solusi'), 
                                        show='headings',
                                        height=10)
        
        conflict_columns = [
            ('Tipe', 120), ('Entitas', 100), ('Hari', 80), 
            ('Waktu', 120), ('Detail1', 180), ('Detail2', 180), ('Solusi', 200)
        ]
        
        for col, width in conflict_columns:
            self.conflict_tree.heading(col, text=col)
            self.conflict_tree.column(col, width=width, anchor='w')
        
        conflict_scroll = ttk.Scrollbar(self.conflict_frame, 
                                      orient="vertical", 
                                      command=self.conflict_tree.yview)
        conflict_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.conflict_tree.configure(yscrollcommand=conflict_scroll.set)
        self.conflict_tree.pack(fill=tk.BOTH, expand=True)
        
        filter_frame = ttk.Frame(self.conflict_frame)
        filter_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(filter_frame, text="Filter Konflik:").pack(side=tk.LEFT)
        
        self.conflict_filter = tk.StringVar(value="Semua")
        filters = [
            ("Semua", "Semua"),
            ("Dosen", "lecturer"),
            ("Ruangan", "room"),
            ("Kapasitas", "capacity"),
            ("Kelas", "class"),
            ("Ruangan Kosong", "empty_room"),
            ("Waktu Istirahat", "break_time")  # Filter baru untuk waktu istirahat
        ]
        
        for text, value in filters:
            ttk.Radiobutton(filter_frame, 
                          text=text, 
                          variable=self.conflict_filter, 
                          value=value,
                          command=self.refresh_conflicts).pack(side=tk.LEFT, padx=5)

    def load_excel_data(self):
        path = filedialog.askopenfilename(title="Pilih File Excel", filetypes=[("Excel Files", "*.xlsx")])
        if path and self.generator.load_data(path):
            self.lecturer_dropdown["values"] = self.generator.lecturers
            if self.generator.lecturers:
                self.lecturer_var.set(self.generator.lecturers[0])
                self.show_lecturer_schedule()
            messagebox.showinfo("Sukses", "Data jadwal berhasil dimuat.")

    def load_room_data_json(self):
        path = filedialog.askopenfilename(title="Pilih File Ruangan (JSON)", filetypes=[("JSON Files", "*.json")])
        if path and self.generator.load_rooms(path):
            messagebox.showinfo("Sukses", "Data ruangan berhasil dimuat.")

    def load_room_data_excel(self):
        path = filedialog.askopenfilename(title="Pilih File Ruangan (Excel)", filetypes=[("Excel Files", "*.xlsx")])
        if path and self.generator.load_rooms_from_excel(path):
            messagebox.showinfo("Sukses", "Data ruangan berhasil dimuat.")

    def show_lecturer_schedule(self, event=None):
        lecturer = self.lecturer_var.get()
        self.schedule_tree.delete(*self.schedule_tree.get_children())
        
        all_schedules = self.generator.fixed_schedules + self.generator.generated_schedules
        filtered_schedules = [s for s in all_schedules if s['dosen'] == lecturer]
        
        if self.current_filter_hari and self.current_filter_hari != 'Semua':
            filtered_schedules = [s for s in filtered_schedules if s['hari'] == self.current_filter_hari]
        
        if self.sort_order_hari == 'asc':
            filtered_schedules.sort(key=lambda x: x['hari'])
        else:
            filtered_schedules.sort(key=lambda x: x['hari'], reverse=True)
        
        for s in filtered_schedules:
            room_capacity = ""
            if s.get('ruangan') and s['ruangan'] != 'Online':
                room_capacity = self.generator.room_capacities.get(s['ruangan'], '?')
            
            self.schedule_tree.insert('', 'end', values=(
                s['hari'],
                s['mata_kuliah'],
                s['kelas'],
                s.get('ruangan', ''),
                room_capacity,
                s['jam'],
                s['sks'],
                s['semester'],
                s['dosen'],
                s.get('jumlah_mahasiswa', '')
            ))

    def apply_filters(self, event=None):
        self.current_filter_hari = self.hari_var.get()
        self.show_lecturer_schedule()

    def toggle_sort_hari(self):
        if self.sort_order_hari == 'asc':
            self.sort_order_hari = 'desc'
            self.sort_hari_btn.config(text="Sort Hari (Z-A)")
        else:
            self.sort_order_hari = 'asc'
            self.sort_hari_btn.config(text="Sort Hari (A-Z)")
        self.show_lecturer_schedule()

    def generate_for_lecturer(self):
        lecturer = self.lecturer_var.get()
        if not lecturer:
            messagebox.showwarning("Peringatan", "Pilih dosen terlebih dahulu!")
            return
        
        if self.generator.generate_schedule_for_lecturer(lecturer):
            self.show_lecturer_schedule()
            messagebox.showinfo("Sukses", f"Jadwal untuk {lecturer} berhasil digenerate.")
        else:
            messagebox.showerror("Gagal", f"Gagal generate jadwal untuk {lecturer}.")

    def clear_rooms(self):
        if self.generator.clear_all_rooms():
            self.show_lecturer_schedule()
            messagebox.showinfo("Sukses", "Semua ruangan berhasil dihapus!")
        else:
            messagebox.showerror("Gagal", "Gagal menghapus ruangan")

    def generate_rooms(self):
        if self.generator.fill_empty_rooms_randomly():
            self.show_lecturer_schedule()
            messagebox.showinfo("Sukses", "Ruangan berhasil diacak ulang!")
        else:
            messagebox.showerror("Gagal", "Gagal mengacak ruangan")

    def save_schedule_all(self):
        all_sched = self.generator.fixed_schedules + self.generator.generated_schedules
        folder = filedialog.askdirectory(title="Pilih Folder Output")
        if folder:
            out = self.generator.save_to_excel(all_sched, "templates/schedule_template.xlsx", folder)
            if out:
                messagebox.showinfo("Sukses", f"Jadwal disimpan di:\n{out}")
                os.startfile(folder)

    def save_schedule_for_current_lecturer(self):
        lecturer = self.lecturer_var.get()
        schedules = [s for s in self.generator.fixed_schedules + self.generator.generated_schedules if s['dosen'] == lecturer]
        if not schedules:
            messagebox.showwarning("Peringatan", "Tidak ada jadwal untuk dosen ini!")
            return
        
        folder = filedialog.askdirectory(title="Pilih Folder Output")
        if folder:
            out = self.generator.save_to_excel(schedules, "templates/schedule_template.xlsx", folder)
            if out:
                messagebox.showinfo("Sukses", f"Jadwal untuk {lecturer} disimpan:\n{out}")
                os.startfile(folder)

    def show_conflicts(self):
        self.conflict_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.refresh_conflicts()
    
    def refresh_conflicts(self):
        self.conflict_tree.delete(*self.conflict_tree.get_children())
        conflicts = self.generator.find_all_conflicts()
        filter_type = self.conflict_filter.get()
        
        conflict_types = {
            'Semua': ['lecturer', 'room', 'class', 'capacity', 'empty_room', 'break_time'],
            'lecturer': ['lecturer'],
            'room': ['room'],
            'capacity': ['capacity'],
            'class': ['class'],
            'empty_room': ['empty_room'],
            'break_time': ['break_time']
        }.get(filter_type, ['lecturer', 'room', 'class', 'capacity', 'empty_room', 'break_time'])
        
        for c_type in conflict_types:
            for conflict in conflicts[c_type]:
                solutions = self.generator.suggest_conflict_resolutions(conflict)
                solution_text = solutions[0] if solutions else "Perlu penyesuaian manual"
                
                if c_type == 'lecturer':
                    self.conflict_tree.insert('', 'end', values=(
                        conflict['conflict_type'],
                        conflict['dosen'],
                        conflict['hari'],
                        conflict['waktu'],
                        f"{conflict['schedule1']['mata_kuliah']} ({conflict['schedule1']['kelas']})",
                        f"{conflict['schedule2']['mata_kuliah']} ({conflict['schedule2']['kelas']})",
                        solution_text
                    ))
                elif c_type == 'room':
                    self.conflict_tree.insert('', 'end', values=(
                        conflict['conflict_type'],
                        conflict['ruangan'],
                        conflict['hari'],
                        conflict['waktu'],
                        f"{conflict['schedule1']['mata_kuliah']} ({conflict['schedule1']['kelas']})",
                        f"{conflict['schedule2']['mata_kuliah']} ({conflict['schedule2']['kelas']})",
                        solution_text
                    ))
                elif c_type == 'class':
                    self.conflict_tree.insert('', 'end', values=(
                        conflict['conflict_type'],
                        conflict['kelas'],
                        conflict['hari'],
                        conflict['waktu'],
                        f"{conflict['schedule1']['mata_kuliah']} (Dosen: {conflict['schedule1']['dosen']})",
                        f"{conflict['schedule2']['mata_kuliah']} (Dosen: {conflict['schedule2']['dosen']})",
                        solution_text
                    ))
                elif c_type == 'capacity':
                    self.conflict_tree.insert('', 'end', values=(
                        conflict['conflict_type'],
                        f"{conflict['ruangan']} (Kap: {conflict['kapasitas']})",
                        conflict['schedule']['hari'],
                        conflict['schedule']['jam'],
                        f"{conflict['schedule']['mata_kuliah']} ({conflict['schedule']['kelas']})",
                        f"Mahasiswa: {conflict['mahasiswa']}",
                        solution_text
                    ))
                elif c_type == 'empty_room':
                    self.conflict_tree.insert('', 'end', values=(
                        conflict['conflict_type'],
                        f"{conflict['ruangan']} (Lt.{conflict['lantai']})",
                        conflict['hari'],
                        conflict['waktu'],
                        f"Kapasitas: {conflict['kapasitas']}",
                        "-",
                        solution_text
                    ))
                elif c_type == 'break_time':
                    self.conflict_tree.insert('', 'end', values=(
                        conflict['conflict_type'],
                        conflict['dosen'],
                        conflict['hari'],
                        conflict['waktu'],
                        f"{conflict['schedule']['mata_kuliah']} ({conflict['schedule']['kelas']})",
                        "Waktu istirahat",
                        solution_text
                    ))

    def show_manual_input(self):
        ManualInputDialog(self.root, self.generator, self.show_lecturer_schedule)

    def on_schedule_select(self, event):
        selected = self.schedule_tree.selection()
        if selected:
            item = self.schedule_tree.item(selected[0])
            values = item['values']
            
            # Find the corresponding schedule in the generator
            lecturer = self.lecturer_var.get()
            all_schedules = self.generator.fixed_schedules + self.generator.generated_schedules
            for s in all_schedules:
                if (s['dosen'] == lecturer and 
                    s['hari'] == values[0] and
                    s['mata_kuliah'] == values[1] and
                    s['kelas'] == values[2] and
                    s.get('ruangan', '') == values[3] and
                    s['jam'] == values[5] and
                    s['sks'] == values[6] and
                    s['semester'] == values[7] and
                    s.get('jumlah_mahasiswa', '') == values[9]):
                    self.selected_schedule = s
                    break
        else:
            self.selected_schedule = None

    def edit_selected_schedule(self):
        if not self.selected_schedule:
            messagebox.showwarning("Peringatan", "Pilih jadwal yang akan diedit!")
            return
            
        # Buka dialog edit untuk semua jenis jadwal
        ManualInputDialog(self.root, self.generator, self.show_lecturer_schedule, self.selected_schedule)

    def delete_selected_schedule(self):
        if not self.selected_schedule:
            messagebox.showwarning("Peringatan", "Pilih jadwal yang akan dihapus!")
            return
            
        if messagebox.askyesno("Konfirmasi", "Apakah Anda yakin ingin menghapus jadwal ini?"):
            if self.selected_schedule.get('source') == 'excel':
                new_schedule = self.selected_schedule.copy()
                new_schedule['hari'] = ""
                new_schedule['jam'] = ""
                new_schedule['ruangan'] = ""
                if self.generator.edit_schedule(self.selected_schedule, new_schedule):
                    messagebox.showinfo("Sukses", "Jadwal dihapus (dikosongkan di file Excel)!")
                    self.show_lecturer_schedule()
                    self.selected_schedule = None
                else:
                    messagebox.showerror("Gagal", "Gagal menghapus jadwal")
            else:
                if self.generator.remove_schedule(self.selected_schedule):
                    messagebox.showinfo("Sukses", "Jadwal berhasil dihapus!")
                    self.show_lecturer_schedule()
                    self.selected_schedule = None
                else:
                    messagebox.showerror("Gagal", "Gagal menghapus jadwal")

    def resolve_conflicts(self):
        resolved = self.generator.auto_resolve_conflicts()
        if resolved > 0:
            messagebox.showinfo("Sukses", f"Berhasil menyelesaikan {resolved} konflik!")
            self.show_lecturer_schedule()
            self.refresh_conflicts()
        else:
            messagebox.showinfo("Info", "Tidak ada konflik yang bisa diselesaikan secara otomatis")
            
    def save_to_original_excel(self):
        """Simpan semua perubahan ke file Excel asli"""
        if not self.generator.excel_path:
            messagebox.showwarning("Peringatan", "Tidak ada file Excel yang dimuat!")
            return
            
        # Kumpulkan semua jadwal yang berasal dari Excel
        excel_schedules = [s for s in self.generator.fixed_schedules if s.get('source') == 'excel']
        
        # Simpan ke file Excel asli
        if self.generator.save_to_excel(excel_schedules, "templates/schedule_template.xlsx", os.path.dirname(self.generator.excel_path)):
            messagebox.showinfo("Sukses", f"Perubahan disimpan ke file Excel asli:\n{self.generator.excel_path}")
        else:
            messagebox.showerror("Gagal", "Gagal menyimpan ke file Excel asli")
    
    def add_break_time(self):
        """Menampilkan dialog untuk menambahkan waktu istirahat dosen"""
        BreakTimeDialog(self.root, self.generator, self.refresh_conflicts)


if __name__ == "__main__":
    root = tk.Tk()
    app = ScheduleApp(root)
    root.mainloop()
